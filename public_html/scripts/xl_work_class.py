import openpyxl as oxl
from openpyxl.styles import Font
from openpyxl.chart import PieChart3D, Reference
import logging
import copy

class Xl_work:
    """Класс для создания неотформатированной версии отчета
        обладает публичными методами:
        open_file, для загрузки Excel файла в openpyxl
        start, для создания структуры итоговой версии отчета
    
    """
    def __init__(self, web_src: str, bit_src: str, done_src: str) -> None:
        """Создает атрибуты класса Xl_work
        Также проверяет и классифицирует ошибки при загрузке фала пользователем, что в дальнейшем
        определяет выводимое на экран сообщение после ответа сервера

        Args:
            web_src (str):  путь к Excel файлу выгрузки с веб-системы, defaults to None
            bit_src (str): путь к Excel файлу выгрузки с Битрикс24, defaults to None
            done_src (str):  путь к итоговому файлу, defaults to None
        """
        self._B24_TITLES = {
            'Название': 'B',
            'Описание': 'C',
            'Бюро': 'U',
            'Статус': 'J'
            }
        self._WEB_TITLES = {
            'Название': 'F'
        }
        self.paths = [web_src, bit_src]
        self.pathDone = done_src
        self.error = ''
        self.message = ''
        self.b24 = dict()
        check1 = self.__correct_file(copy.deepcopy(self.paths[0]))
        check2 = self.__correct_file(copy.deepcopy(self.paths[1]))
        if check1 == 'web' and check2 == 'bitrix':
            self.error = ''
        elif check1 == 'can not read' or check2 == 'can not read':
            if check1 == 'can not read': self.error += 'Файл web не читается'
            if check2 == 'can not read': self.error += 'Файл bitrix не читается'
        elif check1 == check2:
            self.error = 'Файлы одинаковы'
        elif check2 == 'web' and check1 == 'bitrix':
            self.error = 'Файлы перепутаны местами'
        else:
            if check1 == 'unknown':
                self.error += 'Файл web не тот'
            if check2 == 'unknown':
                self.error += 'Файл bitrix не тот'

    def __make_dict_b24(self, sheet) -> dict:
        buros = {}

        for i, el in enumerate(sheet[self._B24_TITLES['Название']]):
            if el.value[:2] == 'ПЭ':
                try:
                    for buro in sheet[self._B24_TITLES['Бюро']+str(i+1)].value.split(', '):
                        buro = buro[5:].capitalize()
                        if buro in buros.keys():
                            buros[buro].append({'name': (el.value[4:], sheet[self._B24_TITLES['Описание']+str(i+1)].value), 'status': sheet[self._B24_TITLES['Статус']+str(i+1)].value != 'Завершена'})
                        else:
                            buros[buro] = list()
                            buros[buro].append({'name': (el.value[4:], sheet[self._B24_TITLES['Описание']+str(i+1)].value), 'status': sheet[self._B24_TITLES['Статус']+str(i+1)].value != 'Завершена'})
                except Exception as e:
                    # self._message(e)
                    pass
        return buros

    def __correct_file(self, path) -> str:
        """Проверяет соответсвие столбцов в файле исходному шаблону.
        Args:
            path (str): путь либо объект проверяемого файла

        Returns:
            str: Возвращает к какому из 3 типов принадлежит загруженный отчетов - из Битрикса
            из Веб-системы или неизвестного происхождения
        """
        wb = self.open_file(path)
        if wb == False:
            return 'can not read'
        else:
            sheet = wb.active
            if sheet[self._B24_TITLES['Название']+'1'].value == 'Название':
                self.b24 = self.__make_dict_b24(sheet)
                wb.close()
                return 'bitrix'
            elif sheet.cell(row=1, column=7).value == 'Опытный узел':
                wb.close()
                return 'web'
            else:
                wb.close()
                return 'unknown'

    def __delete_unwanted_rows(self)->None:
        """Удаляет строки, не содержащие необходимой информации, например
        строки, дублирующие шапку страницы

        :rtype: None
        
        """

        wb = self.open_file(self.paths[1])
        sheet = wb.active

        for i in range(sheet.max_row, 2, -1):
            if (sheet.cell(column=2, row=i).value == 'Название') or (sheet.cell(column=2, row=i).value == None):
                sheet.delete_rows(amount=1, idx=i)

        wb.save(self.paths[1])
        wb.close()
        
    def __create_sheets(self) -> None:
        """ Создает в листы с информацией для каждого бюро в итоговом файле
        При этом, передаются только незавершенные записи о ПЭ 

        :rtype:None

        """
        wb_web = self.open_file(self.paths[0])
        
        ws_web = wb_web.active
        for buro in self.b24.keys():
            wb_web.create_sheet(buro)
        
        wb_web.create_sheet('Конфликты')
        wb_web.save(self.pathDone)
        wb_web.close()
    
    def __count_tasks_in_departments(self)->dict:
        """Создает словарь с названиями всех бюро и количество программ ПЭ в каждом из них

        Returns:
            dict: Словарь с названиями всех бюро и количество программ ПЭ в каждом из них
        """

        amount = {}
        for buro, names in self.b24.items():
            amount[buro] = len(names)//2
        
        return amount

    def __spread_on_sheets(self) -> None:
        """Переносит информацию из веб-системы на лист соответсвующего бюро в итоговом файле

        Args:
            links (dict): словарь ключей-названий бюро
        """
        wb_web = self.open_file(self.paths[0])
        wb_done = self.open_file(self.pathDone)

        ws_web = wb_web.active
        web_names_dict = {}
        for i_row, names in enumerate(ws_web[self._WEB_TITLES['Название']]):

            names = names.value.replace('  ', ' ')
            if i_row == 0:
                continue

            for name in names.split('; '):
                # Формируем строку Excel в массив
                our_row = []
                for i_column, cell in enumerate(ws_web[i_row+1]):
                    if i_column == 5: # Наше название
                        our_row.append(name)
                        continue
                    if i_column == ws_web.max_column-1: # Бюро обрезаем
                        continue
                    our_row.append(cell.value)
                if name in web_names_dict.keys():
                    web_names_dict[name].append(our_row)
                else:
                    web_names_dict[name] = [our_row]
        schet = 0
        for sheet in wb_done.sheetnames[1:-1]:
            for programm in filter(lambda each: each['status'] == True, self.b24[sheet]):
                for name in programm['name']:
                    if name in web_names_dict.keys():
                        for row in web_names_dict[name]:
                            wb_done[sheet].append(row)
                        schet += 1
                        break
                else:
                    # self._message('start')
                    # self._message(list(each for each in programm['name']))
                    # wb_done['Конфликты'].append(our_row)
                    pass
        self._message(schet)
        self._message(len(list(web_names_dict.keys())))

        wb_done.save(self.pathDone)
        wb_done.close()
        wb_web.close()

    def __count_unique(self, column: int, path = None, sheet = None)->int:
        """Рассчитывает число уникальных элементов, получив на вход номер столбца с идентификаторами, а 
        также путь к файлу или лист уже открытого

        Args:
            column (int): Номер колонки, где находятся индексы
            path (_type_, optional): Путь к файлу. Defaults to None.
            sheet (_type_, optional): Лист в открытом файле. Defaults to None.

        Returns:
            int: Число машин
        """
        if sheet:
            unique_elems = []

            for i in range(2, sheet.max_row):
                if (sheet.cell(column=column, row=i).value not in unique_elems):
                    unique_elems.append(sheet.cell(column=column, row=i).value)
            return(len(unique_elems))
        if path:
            wb = self.open_file(path)

            sheet = wb.active

            unique_elems = []

            for i in range(2, sheet.max_row):
                if (sheet.cell(column=column, row=i).value not in unique_elems):
                    unique_elems.append(sheet.cell(column=column, row=i).value)
            wb.save(path)
            wb.close()
            return (len(unique_elems))
        
    def __stat(self)->None:
        """Создает в итоговом файле лист со статистикой
        На этом листе выводит общее количество тракторов, которое находит через функцию __count_number_of_machines,
        список всех бюро с количеством проектов ПЭ в каждом из них на основе словаря,
        полученного из функции __count_tasks_in_departments и создает круговую диграмму распределения проектов ПЭ по бюро

        Returns:
            None
        
        """

        num_of_machines = self.__count_unique(path=self.paths[0], column=2)
        num_of_programms = self.__count_unique(path=self.paths[0], column=6)
        path = self.pathDone
        data = self.__count_tasks_in_departments()

        wb = oxl.load_workbook(filename=path)
        wb.create_sheet('Статистика')
        wb.active=wb['Статистика']
        sheet = wb.active

        sheet.column_dimensions['A'].width = 36.29
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25

        """Выведение числа тракторов"""
        sheet.cell(column=1, row=1).value = 'Количество тракторов'
        sheet.cell(column=1, row=1).font = Font(name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=1).value = num_of_machines
        sheet.cell(column=2, row=1).font = Font(name='Times New Roman', bold=False, size=12)

        '''Выведение общ. числа ПЭ'''
        sheet.cell(column=1, row=2).value = 'Количество программ ПЭ'
        sheet.cell(column=1, row=2).font = Font(name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=2).value = num_of_programms
        sheet.cell(column=2, row=2).font = Font(name='Times New Roman', bold=False, size=12)

        """Таблица с числом проектов у каждого бюро"""
        sheet.cell(column=1, row=4).value = 'Бюро'
        sheet.cell(column=1, row=4).font = Font(name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=4).value = 'Кол-во ПЭ'
        sheet.cell(column=2, row=4).font = Font(name='Times New Roman', bold=True, size=12)
        sheet.cell(column=3, row=4).value = 'Кол-во тракторов в ПЭ'
        sheet.cell(column=3, row=4).font = Font(name='Times New Roman', bold=True, size=12)
        for row_index, (key, value) in enumerate(data.items(), start=5):
            sheet[f'A{row_index}'] = key
            sheet[f'B{row_index}'] = value
            sheet[f'C{row_index}'] = self.__count_unique(column=2, sheet=wb[wb.sheetnames[row_index-4]])

        for i in range(1, len(wb.sheetnames)-2):
            sheet.cell(row = 4+i, column=1).hyperlink = f"#'{wb.sheetnames[i]}'!A1"


        """Создание диаграммы"""
        chart = PieChart3D()
        labels = Reference(sheet, min_col=1, min_row=5, max_row=sheet.max_row, max_col=1)
        info = Reference(sheet, min_col=2, min_row=4, max_row=sheet.max_row, max_col=2)
        chart.add_data(info, titles_from_data=True)
        chart.set_categories(labels)

        chart.width = 15
        chart.height = 12

        chart.legend.position = 'b'

        chart.title = 'Программы ПЭ в бюро'
        chart.series[0].explosion = 10  


        sheet.add_chart(chart, 'E1')

        wb.move_sheet(wb.active, offset=-(len(wb.sheetnames) - 1))
        wb.active = wb['Статистика']
        
        wb.save(self.pathDone)
        wb.close()
    
    def open_file(self, path):
        """Открывает файл

        Args:
            path (_type_): путь к файлу

        Returns:
            _type_: Excel WorkBook
        """

        try:
            wb = oxl.load_workbook(filename=path)
        except:
            return False
        return wb
    
    def _message(self, message: str) -> None:
        message = str(message)
        if message not in self.message:
            self.message += message + '|||'

    def start(self) -> None:
        """ Запускает весь процесс совмещения двух файлов, перед этим выполняя проверку на ошибки при добавлении исходных файлов
        в итогоговом файле распределяяет информацию по листам с бюро и создает лист статистики

        :rtype: None
        
        """
        self.__delete_unwanted_rows()
        self.__create_sheets()
        self.__spread_on_sheets()
        self.__stat()
