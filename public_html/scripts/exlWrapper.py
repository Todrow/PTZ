import openpyxl as oxl
from openpyxl.styles import Font, Alignment

def find_all(a_str, sub): # Все вхождения подстроки в строку
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub)

class ExcelWrapper: # 
    def __init__(self, deleteList: list, blackList: list, path) -> None:   
        # try:
        #     x2x = XLS2XLSX(path)
        #     self.wb = x2x.to_xlsx()
        # except:
        #     self.wb = oxl.load_workbook(filename=path)
        self.wb = self.__open_file(path)
        self.DELETE_LIST = deleteList
        self.ws = self.wb.active
        self.BLACK_LIST = blackList
        self.path = path

    def __open_file(self, path):
        """Открывает файл как work_book в openpyxl

        :param path: путь к файлу, который необходимо открыть
        :type path: str 
        
        """
        wb = oxl.load_workbook(filename=path)
        return wb

    def __deleteColumns(self, blackList: list) -> None: # Удаление из blackList
        title = self.ws[1]
        del_i = 1
        for index, name in enumerate(title):
            if name.value in blackList:
                self.ws.delete_cols(index+del_i, 1)
                del_i -= 1

    def __pasteValues(self, blackList: list) -> None: # Вставить пропущенные значения
        for index, row in enumerate(self.ws.values, start=2):
            for indexColumn, value in enumerate(row, start=1):
                if self.ws.cell(row=1, column=indexColumn).value in blackList:
                    continue
                if self.ws.cell(row=index, column=indexColumn).value is None or self.ws.cell(row=index, column=indexColumn).value == '':
                    self.ws.cell(row=index, column=indexColumn).value = self.ws.cell(row=index-1, column=indexColumn).value

    def formatTitles(self, ws, do_add: bool) -> None:
        """ Форматирует названия столбцов:
        задает им ширину, выравнивание, шрифт и его начертание

        :param ws: рабочий лист Excel

        :param do_add: сообщает о необходимости добавлять шапку
        :type do_add: bool

        :rtype: None
        """
        if do_add:
            ws.insert_rows(0)
        names = ["Модель трактора", "№ трактора", "Граничная дата гарантии", "Продолжительность контроля, м/ч", "Наработка, м/ч", "Опытный узел", "Дата и время обращения", "ПЭ: Комментарий", "Дефект выявлен на м/ч", "Разработчик программы ПЭ"]
        for index, el in enumerate(ws[1]):
            el.font = Font(name="Times New Roman", bold=True, size=12)
            el.value = names[index]
        ws.auto_filter.ref = ws.dimensions

    def formattingCells(self, ws) -> None: 
        """Форматирует ячейки таблицы:
        задает ширину колонок и выравнивание текста

        :param ws:рабочий лист Excel

        :rtype: None        
        """
        ws.row_dimensions[1].height = 30
        widths = {'A': 17.554, 'B': 16.332, 'C': 19.109, 'D': 23.109, 'E': 12.886, 'F': 53.441, 'G': 18.664, 'H': 105.441, 'I': 20.332, 'J': 25.441}
        for row in ws.rows:
            for cell in row:
                if cell.column_letter in widths.keys():
                    ws.column_dimensions[cell.column_letter].width = widths[cell.column_letter]
                    if cell.column_letter not in ['F', 'H', 'J'] or cell.row == 1:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    if cell.column_letter == 'J' and cell.value.count(',') > 1:
                        commas = list(find_all(cell.value, ','))
                        ad = 0
                        for each in commas:
                            each += ad
                            if each+2 < len(cell.value):
                                cell.value = cell.value[:each+2] + '\n' + cell.value[each+2:]
                                ad += 1
    
    def format(self) -> None: # Применяем всё форматирование
        # Удаляем лишнее
        self.__deleteColumns(self.DELETE_LIST)
        # Вставляем пропущенное
        self.__pasteValues(self.BLACK_LIST)
        # Форматируем заголовки
        self.formatTitles(self.ws, False)
        # Форматируем размеры ячеек
        self.formattingCells(self.ws)
        self.save(self.path)

    def save(self, path: str) -> None: # Сохраняем изменнёный файл
        self.wb.save(path)
        self.wb.close()
