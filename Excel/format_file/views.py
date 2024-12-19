from django.shortcuts import render
from scripts.exlWrapper import ExcelWrapper
from scripts.xl_work_class import Xl_work
from django.http import HttpResponse, JsonResponse
import os
import openpyxl as oxl

import openpyxl as oxl
import logging

_logger = logging.getLogger(__name__)

path_done = os.path.abspath('./uploads/')+'/'

def format_file(path_file:str, path_done:str, addAvr) -> str:
    ew = ExcelWrapper(['Вложения', 'Последний раз обновлено', 'Статус', 'Наименование сервисного центра'], ['ПЭ: дата время', 'ПЭ: Комментарий', 'ПЭ: наработка м/ч'], path_file)
    ew.format(addAverage=addAvr)
    wb = oxl.load_workbook(filename=path_file)
    for sheet in wb.sheetnames[2:]:
        ew.formatTitles(wb[sheet], True)
        ew.formattingCells(wb[sheet])
    wb.save(path_done)
    wb.close()
    return

def index_2(request):
    """Выводит страницу с программой форматирования отчета

    Returns:
        _type_: _description_
    """

    global path_done
    try:
        import time
        for filename in os.listdir(path_done):
            f = os.path.join(path_done, filename)
            if time.time() - os.path.getctime(f) > 300:
                os.remove(f)
    except:
        pass

    if request.method == 'POST' and request.FILES:
        file = request.FILES.get('file')
        addAvr = request.POST.get('avr_btn')

        try:
            wb = oxl.load_workbook(file)
        except:
            return JsonResponse({"error": 'Файл не читается', "id": str(request.META['HTTP_ID'])})

        format_file(file, path_done+request.META['HTTP_ID']+'.xlsx', addAvr=addAvr)

        try:
            os.remove(file)
        except:
            pass

        return JsonResponse({"error": '', "id": str(request.META['HTTP_ID'])})
    return render(template_name='index_2.html', request=request)

def download_file2(request, id):
    global path_done
    with open(path_done+id+'.xlsx', 'rb') as file:
        response = HttpResponse(file.read())
        response['Content-Disposition'] = 'attachment; filename=' + 'finished_report.xlsx'
    return response
