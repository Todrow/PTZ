import openpyxl
from xls2xlsx import XLS2XLSX

path = 'Отчет по ПЭ от 02.08.2024.xls'

try:
    x2x = XLS2XLSX(path)
    wb = x2x.to_xlsx()
except:
    wb = openpyxl.load_workbook(filename=path)
    

sheet = wb.active 

unique_machines = []

for i in range(2, sheet.max_row):
    if (sheet.cell(column=2, row=i).value not in unique_machines):
        unique_machines.append(sheet.cell(column=2, row=i).value)

print(len(unique_machines))

wb.close()
