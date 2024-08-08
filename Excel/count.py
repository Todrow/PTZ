import openpyxl

wb = openpyxl.load_workbook('Выгрузка задачи Битрикс.xlsx')
sheet = wb.active
amount = {}

l = sheet.max_row

for i in range(1, l):
    task_name = sheet.cell(column=2, row=i).value
    first_two = task_name[:2]
    if (first_two == 'ПЭ'):
        if (sheet.cell(column=10, row=i).value != 'Завершена'):
                tag = sheet.cell(column=21, row=i).value
                tags = tag.split(', ')
                for each in tags:
                    if (each in amount):
                        amount[each] += 1
                    else:
                        amount[each] = 1

names = list(amount.keys())

for i in range(0, len(names)):
     names[i] = names[i][5:]
  

for i in range(0, len(names)):
    ws = wb.create_sheet(names[i])


wb.save('Выгрузка задачи Битрикс edited.xlsx')
wb.close()  

