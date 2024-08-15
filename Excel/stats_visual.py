import openpyxl as oxl
from openpyxl.styles import Font, Alignment, colors, Color
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList 
from openpyxl.chart.layout import Layout, ManualLayout

file = 'stat test.xlsx'
data = {'key1': 1,
        'key2': 2}

def stat(path: str, num_of_machines: int, data: dict):
    wb = oxl.load_workbook(filename=path)
    sheet = wb.active

    sheet.cell(column=1, row=1).value = 'Количество тракторов'
    sheet.cell(column=1, row=1).font = Font(name='Times New Roman', bold=True, size=12)
    sheet.cell(column=2, row=1).font = Font(name='Times New Roman', bold=False, size=12)
    sheet.column_dimensions['A'].width = 30

    sheet.cell(column=2, row=1).value = num_of_machines

    sheet.append(list(data.keys()))
    sheet.append(list(data.values()))

    
    dll = DataLabelList(showVal=True)
    chart = PieChart()
    labels = Reference(sheet, min_col=1, min_row=2, max_col=len(list(data.keys())), max_row=2)
    info = Reference(sheet, min_col=1, min_row=3, max_col=len(list(data.values())), max_row=3)
    chart.add_data(info, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = ''
    chart.legend.position = 'r'
    chart.title.position = 'r'
    chart.layout=Layout(
        manualLayout=ManualLayout(
            x=0, y=0,
            h=1, w=1,
        )
    )
    

    for i in range(1, len(list(data.keys()))+1):
        sheet.cell(row=2, column=i).font = Font(color=colors.WHITE)
    for i in range(1, len(list(data.keys()))+1):
        sheet.cell(row=3, column=i).font = Font(color=colors.WHITE)
    
    sheet.add_chart(chart, 'D5')
    
    wb.save('stat test.xlsx')
    wb.close()

stat(file, 30, data)
