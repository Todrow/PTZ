import openpyxl as oxl
from openpyxl.styles import Font, Alignment
from xls2xlsx import XLS2XLSX

paths = ["D:/PTZ/Excel/w.xls", "D:/PTZ/Excel/b.xlsx"]



try:
    x2x = XLS2XLSX(paths[1])
    wb2 = x2x.to_xlsx()
except:
    wb2 = oxl.load_workbook(filename=paths[1])

tips = {}
ws2 = wb2.active

for i, el in enumerate(ws2["B"]):
    if el.value[:2] == 'ПЭ':
        tips[el.value[4:]] = ws2['U'+str(i+1)].value.split(', ')


amount = {}

l = ws2.max_row

for i in range(1, l):
    task_name = ws2.cell(column=2, row=i).value
    first_two = task_name[:2]
    if (first_two == 'ПЭ'):
        if (ws2.cell(column=10, row=i).value != 'Завершена'):
                tag = ws2.cell(column=21, row=i).value
                tags = tag.split(', ')
                for each in tags:
                    if (each in amount):
                        amount[each] += 1
                    else:
                        amount[each] = 1

names = list(amount.keys())

for i in range(len(names)):
     names[i] = names[i][5:]
  

try:
    x2x = XLS2XLSX(paths[0])
    wb1 = x2x.to_xlsx()
except:
    wb1 = oxl.load_workbook(filename=paths[0])
ws1 = wb1.active

for i in range(len(names)):
    ws = wb1.create_sheet(names[i])


wb2.close()



first = True
for i, el in enumerate(ws1["F"]): # Не работает
    if first:
        first = False
        continue
    knots = el.value.split('; ')
    # print(knots)
    # print(tips)
    for each in knots:
        if each in tips.keys():
            for byros in tips[each]:
                wb1[byros[5:]].append([cell.value for cell in ws1[i+1]])
        else:
            pass
        # print(ws1[i+1])
wb1.save('./ii.xlsx')
wb1.close()