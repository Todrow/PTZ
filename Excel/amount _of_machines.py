import openpyxl
from xls2xlsx import XLS2XLSX

path1 = 'Отчет по ПЭ от 02.08.2024.xls'
path2 = 'Отчет по ПЭ от 02.08.2024.xlsx'

def count_number_of_machines(path_load, path_save):

    try:
        x2x = XLS2XLSX(path_load)
        wb = x2x.to_xlsx()
    except:
        wb = openpyxl.load_workbook(filename=path_load)
        

    sheet = wb.active 

    unique_machines = []

    for i in range(2, sheet.max_row):
        if (sheet.cell(column=2, row=i).value not in unique_machines):
            unique_machines.append(sheet.cell(column=2, row=i).value)

    print(len(unique_machines))

    wb.save(path_save)
    wb.close()

count_number_of_machines(path1, path2)
