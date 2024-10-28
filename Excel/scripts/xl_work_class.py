import openpyxl as oxl
from openpyxl.styles import Font
from openpyxl.chart import PieChart3D, Reference
import copy

from merge_files.models import ModuleSU, Bureau
from collections import Counter
from openpyxl.styles import Font, Alignment



class Xl_work:
    """Класс для создания неотформатированной версии отчета
        обладает публичными методами:
        open_file, для загрузки Excel файла в openpyxl
        start, для создания структуры итоговой версии отчета

    """

    def __init__(self, web_src: str, bit_src: str, done_src: str) -> None:
        """Создает атрибуты класса Xl_work
        Также проверяет и классифицирует ошибки при загрузке фала пользователем, что в дальнейшем
        определяет выводимое на экран сообщение после ответа сервера

        Args:
            web_src (str):  путь к Excel файлу выгрузки с веб-системы, defaults to None
            bit_src (str): путь к Excel файлу выгрузки с Битрикс24, defaults to None
            done_src (str):  путь к итоговому файлу, defaults to None
        """
        self.paths = [web_src, bit_src]
        self.pathDone = done_src
        self.error = ''
        self.message = ''
        if bit_src is None:
            check1 = self.__correct_file(copy.deepcopy(self.paths[0]))
            check2 = 'bitrix'
        else:
            check1 = self.__correct_file(copy.deepcopy(self.paths[0]))
            check2 = self.__correct_file(copy.deepcopy(self.paths[1]))
        if check1 == 'web' and check2 == 'bitrix':
            self.error = ''
        elif check1 == 'can not read' or check2 == 'can not read':
            if check1 == 'can not read':
                self.error += 'Файл web не читается'
            if check2 == 'can not read':
                self.error += 'Файл bitrix не читается'
        elif check1 == check2:
            self.error = 'Файлы одинаковы'
        elif check2 == 'web' and check1 == 'bitrix':
            self.error = 'Файлы перепутаны местами'
        else:
            if check1 == 'unknown':
                self.error += 'Файл web не тот'
            if check2 == 'unknown':
                self.error += 'Файл bitrix не тот'

    def __correct_file(self, path) -> str:
        """Проверяет соответсвие столбцов в файле исходному шаблону.
        Args:
            path (str): путь либо объект проверяемого файла

        Returns:
            str: Возвращает к какому из 3 типов принадлежит загруженный отчетов - из Битрикса
            из Веб-системы или неизвестного происхождения
        """
        wb = self.open_file(path)
        if wb == False:
            return 'can not read'
        else:
            sheet = wb.active
            if sheet.cell(row=1, column=2).value == 'Название':
                wb.close()
                return 'bitrix'
            elif sheet.cell(row=1, column=7).value == 'Опытный узел':
                wb.close()
                return 'web'
            else:
                wb.close()
                return 'unknown'

    def __delete_unwanted_rows(self) -> None:
        """Удаляет строки, не содержащие необходимой информации, например
        строки, дублирующие шапку страницы

        :rtype: None

        """

        wb = self.open_file(self.paths[1])
        sheet = wb.active

        for i in range(sheet.max_row, 2, -1):
            if (sheet.cell(column=2, row=i).value == 'Название') or (sheet.cell(column=2, row=i).value == None):
                sheet.delete_rows(amount=1, idx=i)

        wb.save(self.paths[1])
        wb.close()

    def __make_link_files(self) -> None:
        """Создет словри из названийи описаний бюро, задействованных в ПЭ

        :rtype: dicts

        """
        wb = self.open_file(self.paths[1])

        if wb:
            ws = wb.active
            for i, el in enumerate(ws["B"], 1):
                if el.value[:2] == 'ПЭ':
                    name = el.value[4:].split('; ')
                    if ws["C"+str(i)].value is not None:
                        disc = ws["C"+str(i)].value
                        if disc[:2] == "ПЭ":
                            name = disc[4:].split('; ')
                    for each in name:
                        bureaus = ws['U'+str(i)].value.split(', ')
                        op_hours_string = ws['AQ'+str(i)].value
                        if op_hours_string:
                            op_hours = str(op_hours_string).split(' ')[0]
                        else:
                            op_hours = 0
                        module_instance, _ = ModuleSU.objects.update_or_create(
                            title=each, defaults={'status': ws["J"+str(i)].value != 'Завершена', 'op_hours': op_hours})
                        for bureau in bureaus:
                            bureau_instance, _ = Bureau.objects.update_or_create(
                                title=bureau)
                            module_instance.bureau_set.add(bureau_instance)
                            bureau_instance.modulesu_set.add(module_instance)

        wb.close()

    def __create_sheets(self) -> None:
        """ Создает в листы с информацией для каждого бюро в итоговом файле
        При этом, передаются только незавершенные записи о ПЭ 

        :rtype:None

        """
        wb_web = self.open_file(self.paths[0])
        bureaus = Bureau.objects.all()

        for bureu in bureaus:
            wb_web.create_sheet(bureu.title[5:].capitalize())

        wb_web.create_sheet('Конфликты')
        wb_web.save(self.pathDone)
        wb_web.close()

    def __count_tasks_in_departments(self) -> dict:
        """Создает словарь с названиями всех бюро и количество программ ПЭ в каждом из них

        Returns:
            dict: Словарь с названиями всех бюро и количество программ ПЭ в каждом из них
        """

        bureaus = Bureau.objects.in_bulk()
        count_modules_in_bureau = {}
        for id in bureaus:
            bureau = bureaus[id]
            count_modules_in_bureau[bureau.title] = len(list(filter(lambda x: x.status, bureau.modules.all())))

        return count_modules_in_bureau

    def __spread_on_sheets(self) -> None:
        """Переносит информацию из веб-системы на лист соответсвующего бюро в итоговом файле

        Args:
            links (dict): словарь ключей-названий бюро
        """
        wb_web = self.open_file(self.paths[0])
        wb_done = self.open_file(self.pathDone)

        linksn = ModuleSU.objects.in_bulk()

        ws_web = wb_web.active
        first = True
        for i, el in enumerate(ws_web["F"]):
            if first:
                first = False
                continue
            knots = el.value.replace('  ', ' ').split('; ')
            for each in knots:
                if each in linksn.keys():
                    if linksn[each].status:
                        for byros in linksn[each].bureaus.all():
                            byros = byros.title
                            our_row = []
                            for j, cell in enumerate(ws_web[i+1]):
                                if j == 5:
                                    our_row.append(each)
                                elif j == 3:
                                    our_row.append(linksn[each].op_hours)
                                elif j == ws_web.max_column-1:
                                    continue
                                else:
                                    our_row.append(cell.value)
                            wb_done[byros[5:].capitalize()].append(our_row)
                else:
                    our_row = []
                    for j, cell in enumerate(ws_web[i+1]):
                        if j == 5:
                            our_row.append(each)
                        elif j == ws_web.max_column:
                            continue
                        else:
                            our_row.append(cell.value)
                    wb_done['Конфликты'].append(our_row)
        wb_done.save(self.pathDone)
        wb_done.close()
        wb_web.close()

    def __count_unique(self, column: int, path=None, sheet=None) -> int:
        """Рассчитывает число уникальных элементов, получив на вход номер столбца с идентификаторами, а 
        также путь к файлу или лист уже открытого

        Args:
            column (int): Номер колонки, где находятся индексы
            path (_type_, optional): Путь к файлу. Defaults to None.
            sheet (_type_, optional): Лист в открытом файле. Defaults to None.

        Returns:
            int: Число машин
        """
        if sheet:
            unique_elems = []

            for i in range(1, sheet.max_row+1):
                if (sheet.cell(column=column, row=i).value not in unique_elems):
                    unique_elems.append(sheet.cell(column=column, row=i).value)
            return (len(unique_elems))
        if path:
            wb = self.open_file(path)

            sheet = wb.active

            unique_elems = []

            for i in range(1, sheet.max_row+1):
                if (sheet.cell(column=column, row=i).value not in unique_elems):
                    unique_elems.append(sheet.cell(column=column, row=i).value)
            wb.save(path)
            wb.close()
            return (len(unique_elems))
    
    def __module_tractor(self, sheet, column_modules: str, column_tractors: str) -> dict:
        module_tractor_dict = dict()
        for i, each in enumerate(sheet[column_modules], 1):
            if each.value in module_tractor_dict.keys():
                module_tractor_dict[each.value].append(sheet[column_tractors+str(i)].value)
            else:
                module_tractor_dict[each.value] = [
                    sheet[column_tractors+str(i)].value]
        for key in module_tractor_dict.keys():
            module_tractor_dict[key] = len(set(module_tractor_dict[key]))
        return module_tractor_dict
    
    def __stat(self) -> None:
        """Создает в итоговом файле лист со статистикой
        На этом листе выводит общее количество тракторов, которое находит через функцию __count_number_of_machines,
        список всех бюро с количеством проектов ПЭ в каждом из них на основе словаря,
        полученного из функции __count_tasks_in_departments и создает круговую диграмму распределения проектов ПЭ по бюро

        Returns:
            None

        """

        num_of_machines = self.__count_unique(path=self.paths[0], column=2)
        num_of_programms = ModuleSU.objects.filter(status=True).count()
        path = self.pathDone
        data = self.__count_tasks_in_departments()

        wb = oxl.load_workbook(filename=path)
        wb.create_sheet('Статистика')
        wb.active = wb['Статистика']
        sheet = wb.active

        sheet.column_dimensions['A'].width = 36.29
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25

        """Выведение числа тракторов"""
        sheet.cell(column=1, row=1).value = 'Количество тракторов'
        sheet.cell(column=1, row=1).font = Font(
            name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=1).value = num_of_machines
        sheet.cell(column=2, row=1).font = Font(
            name='Times New Roman', bold=False, size=12)

        '''Выведение общ. числа ПЭ'''
        sheet.cell(column=1, row=2).value = 'Количество программ ПЭ'
        sheet.cell(column=1, row=2).font = Font(
            name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=2).value = num_of_programms
        sheet.cell(column=2, row=2).font = Font(
            name='Times New Roman', bold=False, size=12)

        """Таблица с числом проектов у каждого бюро"""
        sheet.cell(column=1, row=4).value = 'Бюро'
        sheet.cell(column=1, row=4).font = Font(
            name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=4).value = 'Кол-во ПЭ'
        sheet.cell(column=2, row=4).font = Font(
            name='Times New Roman', bold=True, size=12)
        sheet.cell(column=3, row=4).value = 'Кол-во тракторов в ПЭ'
        sheet.cell(column=3, row=4).font = Font(
            name='Times New Roman', bold=True, size=12)
        for row_index, (key, value) in enumerate(data.items(), start=5):
            sheet[f'A{row_index}'] = key
            sheet[f'B{row_index}'] = value
            sheet[f'C{row_index}'] = self.__count_unique(
                column=2, sheet=wb[wb.sheetnames[row_index-4]])

        for i in range(1, len(wb.sheetnames)-2):
            sheet.cell(
                row=4+i, column=1).hyperlink = f"#'{wb.sheetnames[i]}'!A1"

        """Создание диаграммы"""
        chart = PieChart3D()
        labels = Reference(sheet, min_col=1, min_row=5,
                           max_row=sheet.max_row, max_col=1)
        info = Reference(sheet, min_col=2, min_row=4,
                         max_row=sheet.max_row, max_col=2)
        chart.add_data(info, titles_from_data=True)
        chart.set_categories(labels)

        chart.width = 15
        chart.height = 12

        chart.legend.position = 'b'

        chart.title = 'Программы ПЭ в бюро'
        chart.series[0].explosion = 10

        sheet.add_chart(chart, 'E1')

        wb.move_sheet(wb.active, offset=-(len(wb.sheetnames) - 1))
        wb.active = wb['Статистика']

        wb.save(self.pathDone)
        wb.close()

    def open_file(self, path):
        """Открывает файл

        Args:
            path (_type_): путь к файлу

        Returns:
            _type_: Excel WorkBook
        """

        try:
            wb = oxl.load_workbook(filename=path)
        except Exception as err:
            self._message('error open file:' + str(err))
            return False
        return wb

    def _message(self, message) -> None:
        message = str(message)
        if message not in self.message:
            self.message += message + '|||'

    def __module_tractor(self, sheet, column_modules: str, column_tractors: str) -> dict:
        module_tractor_dict = dict()
        for i, each in enumerate(sheet[column_modules], 1):
            if i == 1:
                continue
            if each.value in module_tractor_dict.keys():
                module_tractor_dict[each.value].append(sheet[column_tractors+str(i)].value)
            else:
                module_tractor_dict[each.value] = [
                    sheet[column_tractors+str(i)].value]
        for key in module_tractor_dict.keys():
            module_tractor_dict[key] = len(set(module_tractor_dict[key]))
        return module_tractor_dict

    def sheet_sort_rows(ws, row_start, row_end=0, cols=None, sorter=None, reverse=False):

        """ Сортирует строки на листе
            Args:
                ws*:          Лист
                row_start(int)*:  Первая строка, которую надо отсортировать
                row_end(int):     Последняя строка, которую надо отсортировать (по умолчанию - последняя строка в листе)
                cols(list):        Колонки, которые надо отсортировать (list) e.g. [2, 4]
                sorter(func):      Функция сортировки (для метода sorted)
                reverse(bool):     Обратная сортировка (развернуть)
            * - Обязательные аргументы
        """

        bottom = ws.max_row
        if row_end == 0:
            row_end = ws.max_row
        right = oxl.get_column_letter(ws.max_column)
        if cols is None:
            cols = range(1, ws.max_column+1)

        array = {}
        for row in range(row_start, row_end+1):
            key = []
            for col in cols:
                key.append(ws.cell(row, col).value)
            array[key] = array.get(key, set()).union({row})

        order = sorted(array, key=sorter, reverse=reverse)

        ws.move_range(f"A{row_start}:{right}{row_end}", bottom)
        dest = row_start
        for src_key in order:
            for row in array[src_key]:
                src = row + bottom
                dist = dest - src
                ws.move_range(f"A{src}:{right}{src}", dist)
                dest += 1
        
    def department_stat(self):
        wb = self.open_file(self.pathDone)


        for each in wb.sheetnames[2:]:
            wb.active = wb[each]
            ws = wb.active
            count = self.__module_tractor(sheet=ws, column_modules='F', column_tractors='B')
            ws.insert_rows(1)
            ws.cell(row=1, column=1, value='Название узла').font = Font(name="Times New Roman", bold=True, size=12)
            ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.cell(row=1, column=6, value='Количество тракторов в программе').font = Font(name="Times New Roman", bold=True, size=12)
            ws.cell(row=1, column=6).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)


            for i, each in enumerate(count.keys()):
                ws.insert_rows(2)
                ws.merge_cells(start_row=i+2, start_column=1, end_row=i+2, end_column=5)
                ws.cell(row=2, column=1, value=each).alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')          
                ws.cell(row=2, column=6, value=count[each]).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                
                
            l = len(count.keys())+2
            ws.insert_rows(l)
            ws.row_dimensions[l].height = 30
            

            for i, each in enumerate(ws['A']):
                if each.value == 'Модель трактора':
                    filter_start = 'A' + str(i+1)

            filter_finish = str(ws.dimensions.split(':')[1])

            filter_dim = filter_start + ':' + filter_finish
            ws.auto_filter.ref = filter_dim

            ws.row_dimensions[1].height = 30
            

        wb.save(self.pathDone)
        wb.close()

            

    def start(self) -> None:
        """ Запускает весь процесс совмещения двух файлов, перед этим выполняя проверку на ошибки при добавлении исходных файлов
        в итогоговом файле распределяяет информацию по листам с бюро и создает лист статистики

        :rtype: None

        """
        if self.paths[1] is not None:
            self.__delete_unwanted_rows()
            self.__make_link_files()
        self.__create_sheets()
        self.__spread_on_sheets()
        self.__stat()
        
