import openpyxl as oxl
from openpyxl.styles import Font, Alignment

def find_all(a_str, sub): # Все вхождения подстроки в строку
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub)

class ExcelWrapper: # 
    def __init__(self, deleteList: list, blackList: list, path) -> None:   
        """Функция innit класса

        Args:
            deleteList (list): Список колонок под удаление
            blackList (list): Список колонок, где не нужно заполнять пропущенные значения
            path (_type_): Полное имя, обрабатываемого файла
        """
        self.wb = self.__open_file(path)
        self.DELETE_LIST = deleteList
        self.ws = self.wb.active
        self.BLACK_LIST = blackList
        self.path = path

    def __open_file(self, path):
        """Открывает файл как work_book в openpyxl

        Args:
            path (_type_): путь к файлу, который необходимо открыть

        Returns:
            _type_: Книга Excel
        """

        wb = oxl.load_workbook(filename=path)
        return wb

    def __deleteColumns(self, blackList: list) -> None:
        """Удаляет заданные колонки

        Args:
            blackList (list): Список колонок под удаление
        """
        title = self.ws[1]
        del_i = 1
        for index, name in enumerate(title):
            if name.value in blackList:
                self.ws.delete_cols(index+del_i, 1)
                del_i -= 1

    def __pasteValues(self, blackList: list) -> None:
        """Вставляет пропущенные значения,
        Берет значения из соседних строк в пустые ячейки, кроме ячеек из
        стобцов в blacklist
        В этих столбцах ничего не добавляется

        Args:
            blackList (list): Список столбцов, где не надо заполнят пропущенные значения
        """
        for index, row in enumerate(self.ws.values, start=2):
            for indexColumn, value in enumerate(row, start=1):
                if self.ws.cell(row=1, column=indexColumn).value in blackList:
                    continue
                if self.ws.cell(row=index, column=indexColumn).value is None or self.ws.cell(row=index, column=indexColumn).value == '':
                    self.ws.cell(row=index, column=indexColumn).value = self.ws.cell(row=index-1, column=indexColumn).value

    def addAverageTime(self, ws):
        tracktors = list()
        total_sum = 0
        for index, row in enumerate(ws.rows, start=0):
            if index == 0:
                continue
            if row[1].value not in tracktors:
                tracktors.append(row[1].value)
                total_sum += int(row[4].value)
        ws[f'D{ws.max_row+1}'] = 'Средняя наработка'
        ws[f'E{ws.max_row}'] = str(total_sum//len(tracktors))

    def formatTitles(self, ws, do_add: bool) -> None:
        """Форматирует названия столбцов:
        задает им ширину, выравнивание, шрифт и его начертание

        Args:
            ws (_type_): рабочий лист Excel
            do_add (bool): сообщает о необходимости добавлять шапку
        """

        

        if do_add:
            ws.insert_rows(0)
        names = ["Модель трактора", "№ трактора", "Граничная дата гарантии", "Продолжительность контроля, м/ч", "Наработка, м/ч", "Опытный узел", "Дата и время обращения", "ПЭ: Комментарий", "Дефект выявлен на м/ч", "Разработчик программы ПЭ"]
        for index, el in enumerate(ws[1]):
            el.font = Font(name="Times New Roman", bold=True, size=12)
            el.value = names[index]

        
        

    def formattingCells(self, ws) -> None: 
        """Форматирует ячейки таблицы:
        задает ширину колонок и выравнивание текста

        Args:
            ws (_type_): рабочий лист Excel
        """

        ws.row_dimensions[1].height = 30
        widths = {'A': 17, 'B': 20, 'C': 19.109, 'D': 23.109, 'E': 12.886, 'F': 53.441, 'G': 18.664, 'H': 105.441, 'I': 20.332, 'J': 25.441}
        for row in ws.rows:
            for cell in row:
                if cell.column_letter in widths.keys():
                    ws.column_dimensions[cell.column_letter].width = widths[cell.column_letter]
                    if cell.column_letter not in ['F', 'H', 'J'] or cell.row == 1:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    # if cell.column_letter == 'J' and cell.value.count(',') > 1:
                    #     commas = list(find_all(cell.value, ','))
                    #     ad = 0
                    #     for each in commas:
                    #         each += ad
                    #         if each+2 < len(cell.value):
                    #             cell.value = cell.value[:each+2] + '\n' + cell.value[each+2:]
                    #             ad += 1
    
    def format(self, addAverage) -> None:
        """Применяет все форматирование
        """
        # Удаляем лишнее
        self.__deleteColumns(self.DELETE_LIST)
        # Вставляем пропущенное
        self.__pasteValues(self.BLACK_LIST)
        # Форматируем заголовки
        self.formatTitles(self.ws, False)
        # Форматируем размеры ячеек
        self.formattingCells(self.ws)
        if addAverage == "true":
            self.addAverageTime(self.ws)
        self.save(self.path)

    def save(self, path: str) -> None:
        """Созраняет файл книгу Excel

        Args:
            path (str): Полное имя файла при сохранении
        """
        self.wb.save(path)
        self.wb.close()
