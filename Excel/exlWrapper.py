import openpyxl as oxl
from openpyxl.styles import Font, Alignment
from xls2xlsx import XLS2XLSX

def find_all(a_str, sub): # Все вхождения подстроки в строку
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub)

class ExcelWrapper: # 
    def __init__(self) -> None:   
        pass

    def __deleteColumns(self, blackList: list) -> None: # Удаление из blackList
        title = list(self.ws.values)[0]
        del_i = 1
        for index, name in enumerate(title):
            if name in blackList:
                self.ws.delete_cols(index+del_i, 1)
                del_i -= 1

    def __pasteValues(self, blackList: list) -> None: # Вставить пропущенные значения
        for index, row in enumerate(self.ws.values, start=2):
            for indexColumn, value in enumerate(row, start=1):
                if self.ws.cell(row=1, column=indexColumn).value in blackList:
                    continue
                if self.ws.cell(row=index, column=indexColumn).value is None or self.ws.cell(row=index, column=indexColumn).value == '':
                    self.ws.cell(row=index, column=indexColumn).value = self.ws.cell(row=index-1, column=indexColumn).value

    def __formatTitles(self) -> None: # Форматирование заголовков
        names = ["Модель трактора", "№ трактора", "Граничная дата гарантии", "Продолжительность контроля, м/ч", "Наработка, м/ч", "Опытный узел", "Дата и время обращения", "ПЭ: Комментарий", "Дефект выявлен на м/ч", "Разработчик программы ПЭ"]
        for index, el in enumerate(self.ws[1]):
            el.font = Font(name="Times New Roman", bold=True, size=12)
            el.value = names[index]
        self.ws.auto_filter.ref = self.ws.dimensions

    def __formattingCells(self) -> None: # Форматируем размеры ячеек
        self.ws.row_dimensions[1].height = 30
        widths = {'A': 17.554, 'B': 16.332, 'C': 19.109, 'D': 23.109, 'E': 12.886, 'F': 53.441, 'G': 18.664, 'H': 105.441, 'I': 20.332, 'J': 25.441}
        for row in self.ws.rows:
            for cell in row:
                if cell.column_letter in widths.keys():
                    self.ws.column_dimensions[cell.column_letter].width = widths[cell.column_letter]
                    if cell.column_letter not in ['F', 'H', 'J'] or cell.row == 1:
                        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')
                    if cell.column_letter == 'J' and cell.value.count(',') > 1:
                        commas = list(find_all(cell.value, ','))
                        ad = 0
                        for each in commas:
                            each += ad
                            if each+2 < len(cell.value):
                                cell.value = cell.value[:each+2] + '\n' + cell.value[each+2:]
                                ad += 1

    def feed(self, deleteList: list, blackList: list, path: str) -> None:
        try:
            x2x = XLS2XLSX(path)
            self.wb = x2x.to_xlsx()
        except:
            self.wb = oxl.load_workbook(filename=path)
        self.DELETE_LIST = deleteList
        self.ws = self.wb.active
        self.BLACK_LIST = blackList
    
    def format(self) -> None: # Применяем всё форматирование
        # Удаляем лишнее
        self.__deleteColumns(self.DELETE_LIST)
        # Вставляем пропущенное
        self.__pasteValues(self.BLACK_LIST)
        # Форматируем заголовки
        self.__formatTitles()
        # Форматируем размеры ячеек
        self.__formattingCells()

    def save(self, path: str) -> None: # Сохраняем изменнёный файл
        self.wb.save(path)
        self.wb.close()