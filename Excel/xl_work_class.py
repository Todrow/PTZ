import openpyxl as oxl
from xls2xlsx import XLS2XLSX
from openpyxl.styles import Font, Alignment, colors, Color, Border, Side
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList 
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
import os

class Xl_work:
    def __init__(self, web_src: str, bit_src: str) -> None:
        self.paths = [web_src, bit_src]
        self.pathDone = r'C:\Users\Aleksandr\Documents\Work\Excel\Report.xlsx'
    
    def __check_format(self)->bool:
        valid_extension = ('.xlsx', '.xls')
        file1 = self.paths[0]
        file2 = self.paths[1]

        if not (file1.lower().endswith(valid_extension) and file2.lower().endswith(valid_extension)):
            return False
    
        return True

    def __make_link_files(self) -> dict:
        try:
            x2x = XLS2XLSX(self.paths[1])
            wb = x2x.to_xlsx()
        except:
            wb = oxl.load_workbook(filename=self.paths[1])
        
        names = {}
        ws = wb.active

        for i, el in enumerate(ws["B"]):
            if el.value[:2] == 'ПЭ':
                names[el.value[4:]] = ws['U'+str(i+1)].value.split(', ')
        wb.close()
        return names
        
    def __create_sheets(self) -> None:
        try:
            try:
                x2x = XLS2XLSX(self.paths[1])
                wb_bit = x2x.to_xlsx()
                print('DoneTry1')
            except:
                wb_bit = oxl.load_workbook(filename=self.paths[1])

            try:
                x2x = XLS2XLSX(self.paths[0])
                wb_web = x2x.to_xlsx()
            except:
                wb_web = oxl.load_workbook(filename=self.paths[0])

            ws_web = wb_web.active
            ws_bit = wb_bit.active

            for i in range(1, ws_bit.max_row):
                task = ws_bit.cell(column=2, row=i).value
                if task[:2] == 'ПЭ' and ws_bit.cell(column=10, row=i).value != 'Завершена':
                    tags = ws_bit.cell(column=21, row=i).value.split(', ')
                    for each in tags:
                        each = each[5:].capitalize()
                        if each in wb_web.sheetnames:
                            continue
                        else: wb_web.create_sheet(each)
            wb_web.create_sheet('Мусорка')
            wb_web.save(self.pathDone)
            wb_web.close()
            wb_bit.close()
        except:
            print('Исходный файл отчета из Битрикс отличается от заданого шаблона')
    
    def __count_tasks_in_departments(self)->dict:

        wb = oxl.load_workbook(self.paths[1])
        sheet = wb.active
        amount = {}
        for i in range(1, sheet.max_row):
            task_name = sheet.cell(column=2, row=i).value
            first_two = task_name[:2]
            if (first_two == 'ПЭ'):
                if (sheet.cell(column=10, row=i).value != 'Завершена'):
                        tag = sheet.cell(column=21, row=i).value
                        tags = tag.split(', ')
                        for each in tags:
                            if (each in amount):
                                amount[each] += 1
                            else:
                                amount[each] = 1
        
        return amount

    def __spread_on_sheets(self, links: dict) -> None:
        try:
            wb_web = self.open_file(self.paths[0])
            wb_done = self.open_file(self.pathDone)

            ws_web = wb_web.active
            first = True
            for i, el in enumerate(ws_web["F"]):
                if first:
                    first = False
                    continue
                knots = el.value.split('; ')
                for each in knots:
                    if each in links.keys():
                        for byros in links[each]:
                            our_row = []
                            for j, cell in enumerate(ws_web[i+1]):
                                if j == 5:
                                    our_row.append(each)
                                elif j == ws_web.max_column-1:
                                    continue
                                else:
                                    our_row.append(cell.value)
                            wb_done[byros[5:].capitalize()].append(our_row)
                    else:
                        wb_done['Мусорка'].append([cell.value for cell in ws_web[i+1]])
            wb_done.save(self.pathDone)
            wb_done.close()
            wb_web.close()
        except:
            print('Исходный файл из веб-системы отличается от заданого шаблона')
            
    def __count_number_of_machines(self, path)->int:

        try:
            x2x = XLS2XLSX(path)
            wb = x2x.to_xlsx()
        except:
            wb = oxl.load_workbook(filename=path)
            

        sheet = wb.active 

        unique_machines = []

        for i in range(2, sheet.max_row):
            if (sheet.cell(column=2, row=i).value not in unique_machines):
                unique_machines.append(sheet.cell(column=2, row=i).value)

        wb.save(path)
        wb.close()
        return (len(unique_machines))

    def __stat(self)->None:
        num_of_machines = self.__count_number_of_machines(self.paths[0])
        path = self.pathDone
        data = self.__count_tasks_in_departments()

        wb = oxl.load_workbook(filename=path)
        wb.create_sheet('Статистика')
        wb.active=wb['Статистика']
        sheet = wb.active

        sheet.cell(column=1, row=1).value = 'Количество тракторов'
        sheet.cell(column=1, row=1).font = Font(name='Times New Roman', bold=True, size=12)
        sheet.cell(column=2, row=1).font = Font(name='Times New Roman', bold=False, size=12)
        sheet.column_dimensions['A'].width = 30

        sheet.cell(column=2, row=1).value = num_of_machines

        for row_index, (key, value) in enumerate(data.items(), start=3):
            sheet[f'A{row_index}'] = key
            sheet[f'B{row_index}'] = value

        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        dll = DataLabelList(showVal=True)
        chart = PieChart()
        labels = Reference(sheet, min_col=1, min_row=3, max_row=sheet.max_row, max_col=1)
        info = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row, max_col=2)
        chart.add_data(info, titles_from_data=True)
        chart.set_categories(labels)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.width = 20
        chart.height = 10
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showPercentage = False
        data_label_font = Font(size=14)
        chart.dataLabels.font = data_label_font
        chart.legend = None


        
        sheet.add_chart(chart, 'D5')
        
        wb.save(self.pathDone)
        wb.close()
    
    def open_file(self, path):
        try:
            x2x = XLS2XLSX(path)
            wb = x2x.to_xlsx()
        except:
            wb = oxl.load_workbook(filename=path)
        return wb

    def start(self) -> None:
        if self.__check_format() == False:
            print('Недопустимый формат файла')
        else:
            self.__create_sheets()
            self.__spread_on_sheets(self.__make_link_files())
            self.__stat()
